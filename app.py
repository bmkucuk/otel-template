#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Otel Yönetim Sistemi — Template (SQLite sürümü)"""
import os
import config_loader as cfg
import smtplib
import tempfile
from datetime import date, datetime
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from zoneinfo import ZoneInfo
from flask import Flask, render_template, request, jsonify, redirect, url_for, session, send_file
import hashlib
from functools import wraps
import database as db
from muhasebe_routes import muh
import muhasebe_db as mdb
from apscheduler.schedulers.background import BackgroundScheduler

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'otelleo2026x!9k')

# Sunucu (Render) UTC'de çalışabilir; otelin gerçek "bugün"ü her zaman Türkiye saatine göre hesaplanır.
TR_TZ = ZoneInfo('Europe/Istanbul')
def bugun():
    return datetime.now(TR_TZ).date()

_TR_AYLAR = ['Ocak','Şubat','Mart','Nisan','Mayıs','Haziran','Temmuz','Ağustos','Eylül','Ekim','Kasım','Aralık']
_TR_GUNLER = ['Pazartesi','Salı','Çarşamba','Perşembe','Cuma','Cumartesi','Pazar']
def tr_tarih(d):
    """Locale'e bağımlı olmadan Türkçe tarih metni üretir: '20 Haziran 2026, Cumartesi'"""
    return f"{d.day} {_TR_AYLAR[d.month-1]} {d.year}, {_TR_GUNLER[d.weekday()]}"

# Kanal (Rez. Formu) → acente kodu eşleştirmesi
KANAL_MAP = {
    'Booking': 'BKG', 'BKG': 'BKG',
    'EXP': 'EXP', 'Expedia': 'EXP',
    'JLY': 'JLY', 'JollyTur': 'JLY',
    'TTS': 'TTS', 'TatilSepeti': 'TTS',
    'ETS': 'ETS', 'ETSTUR': 'ETS', 'ETStur': 'ETS',
    'Telefon': 'Telefon',
    'Kapıdan': 'Kapidan',
}
# Bu acenteler için rez. bedeli direkt müşteriden değil acenteden tahsil edildiği
# için rezervasyon kaydedilirken otomatik muhasebeleşir (borç acente / alacak müşteri
# + komisyon otomatik düşülür). Diğer kanallar (EXP, Telefon, Kapıdan) eskisi gibi kalır.
ACENTE_OTO_KODLAR = {'JLY', 'BKG', 'ETS', 'TTS', 'EXP'}
ACENTE_HESAP_KODU = {'BKG': '320-1', 'EXP': '320-2', 'JLY': '320-3', 'TTS': '320-4', 'ETS': '320-5'}
ACENTE_ADI = {'BKG': 'Booking', 'EXP': 'Expedia', 'JLY': 'JollyTur', 'TTS': 'TatilSepeti', 'ETS': 'ETSTUR'}

# Kullanıcılar artık veritabanında (kullanicilar tablosu) — bkz. database.py init_db()

# ── Aktif Kullanıcı Takibi (tek worker, bellek içi) ─────────────────────────
AKTIF_KULLANICILAR = {}  # {username: {'role':..., 'giris':datetime, 'son_aktivite':datetime}}
AKTIF_ESIK_DK = 10  # bu süre içinde istek atmayan kullanıcı "aktif" sayılmaz

@app.before_request
def _lisans_kontrol():
    acik_yollar = ['/login', '/kurulum', '/static', '/telegram-webhook', '/askida', '/demo-bitti', '/sadmin']
    if not any(request.path.startswith(y) for y in acik_yollar):
        # Kurulum tamamlanmamışsa sihirbaza yönlendir
        if cfg.get('otel.ad', 'Otel Adı') == 'Otel Adı':
            return redirect('/kurulum')
        durum = cfg.lisans_durumu()
        if durum == 'askida':
            return redirect('/askida')
        elif durum == 'demo_bitti':
            return redirect('/demo-bitti')

@app.before_request
def _aktivite_guncelle():
    u = session.get('user')
    if u:
        simdi = datetime.now(TR_TZ)
        kayit = AKTIF_KULLANICILAR.get(u)
        if kayit:
            kayit['son_aktivite'] = simdi
        else:
            AKTIF_KULLANICILAR[u] = {
                'role': session.get('role'),
                'giris': simdi,
                'son_aktivite': simdi,
            }

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('user'):
            return redirect('/login')
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('user'):
            return redirect('/login')
        if session.get('role') not in ('admin', 'partner'):
            return redirect('/')
        return f(*args, **kwargs)
    return decorated

# Muhasebe blueprint
app.register_blueprint(muh)

# ── Template context processor ───────────────────────────────────────────────
@app.context_processor
def inject_lisans():
    def _lisans():
        c = cfg.load_config()
        return {
            'durum':    cfg.lisans_durumu(),
            'kalan':    cfg.demo_kalan_gun(),
            'otel':     c.get('otel', {}),
            'ortaklar': c.get('ortaklar', [])
        }
    return {
        'lisans_bilgi': _lisans,
        'tema_mod': cfg.tema_mod()
    }

# DB başlat
db.init_db()
mdb.init_db()

# ── Auth ─────────────────────────────────────────────────────────────────────

@app.route('/login', methods=['GET','POST'])
def login():
    if session.get('user'):
        return redirect('/')
    if request.method == 'POST':
        username = request.form.get('username','').strip()
        password = request.form.get('password','')
        h = hashlib.sha256(password.encode()).hexdigest()
        user = db.kullanici_getir(username)
        if user and user['hash'] == h:
            session['user'] = username
            session['role'] = user['role']
            session['ad'] = user['ad']
            db.log_yaz(username, user['role'], datetime.now(TR_TZ).strftime('%Y-%m-%d %H:%M:%S'),
                       'giris', 'auth', '/login', f"{username} giriş yaptı", 200)
            return redirect('/')
        db.log_yaz(username or '(boş)', None, datetime.now(TR_TZ).strftime('%Y-%m-%d %H:%M:%S'),
                   'giris_basarisiz', 'auth', '/login', f"{username} için başarısız giriş denemesi", 401)
        return render_template('login.html', error='Kullanıcı adı veya şifre yanlış')
    return render_template('login.html', error=None)

@app.route('/logout')
def logout():
    u, r = session.get('user'), session.get('role')
    if u:
        db.log_yaz(u, r, datetime.now(TR_TZ).strftime('%Y-%m-%d %H:%M:%S'),
                   'cikis', 'auth', '/logout', f"{u} çıkış yaptı", 200)
    AKTIF_KULLANICILAR.pop(session.get('user'), None)
    session.clear()
    return redirect('/login')

@app.route('/sifre-degistir', methods=['GET', 'POST'])
@login_required
def sifre_degistir():
    if session.get('role') not in ('admin', 'partner'):
        return redirect('/')
    hata, basari = None, None
    if request.method == 'POST':
        mevcut = request.form.get('mevcut_sifre', '')
        yeni = request.form.get('yeni_sifre', '')
        tekrar = request.form.get('yeni_sifre_tekrar', '')
        user = db.kullanici_getir(session['user'])
        if not user or user['hash'] != hashlib.sha256(mevcut.encode()).hexdigest():
            hata = 'Mevcut şifre yanlış'
        elif len(yeni) < 6:
            hata = 'Yeni şifre en az 6 karakter olmalı'
        elif yeni != tekrar:
            hata = 'Yeni şifreler birbiriyle eşleşmiyor'
        else:
            db.kullanici_sifre_guncelle(session['user'], hashlib.sha256(yeni.encode()).hexdigest())
            db.log_yaz(session['user'], session.get('role'), datetime.now(TR_TZ).strftime('%Y-%m-%d %H:%M:%S'),
                       'sifre_degistir', 'auth', '/sifre-degistir', f"{session['user']} şifresini değiştirdi", 200)
            basari = 'Şifreniz başarıyla güncellendi'
    return render_template('sifre_degistir.html', hata=hata, basari=basari)

@app.route('/api/aktif-kullanicilar')
@admin_required
def api_aktif_kullanicilar():
    simdi = datetime.now(TR_TZ)
    sonuc = []
    silinecek = []
    for kullanici, bilgi in AKTIF_KULLANICILAR.items():
        fark_dk = (simdi - bilgi['son_aktivite']).total_seconds() / 60
        if fark_dk > AKTIF_ESIK_DK:
            silinecek.append(kullanici)
            continue
        if bilgi['role'] == 'admin':
            continue  # admin widget'ta görünmez
        sonuc.append({
            'kullanici': kullanici,
            'giris': bilgi['giris'].strftime('%H:%M'),
            'son_aktivite_dk': round(fark_dk),
        })
    for k in silinecek:
        AKTIF_KULLANICILAR.pop(k, None)
    return jsonify(sonuc)

@app.route('/islem-gecmisi')
@admin_required
def islem_gecmisi():
    return render_template('islem_gecmisi.html')

@app.route('/api/islem-loglari')
@admin_required
def api_islem_loglari():
    kullanici = request.args.get('kullanici') or None
    baslangic = request.args.get('baslangic') or None
    bitis     = request.args.get('bitis') or None
    arama     = request.args.get('arama') or None
    loglar    = db.log_listele(limit=500, kullanici=kullanici, baslangic=baslangic, bitis=bitis, arama=arama)
    return jsonify(loglar)

@app.route('/api/kullanici-listesi')
@admin_required
def api_kullanici_listesi():
    conn = db.get_conn()
    rows = conn.execute(
        "SELECT username FROM kullanicilar WHERE aktif=1 ORDER BY username"
    ).fetchall()
    conn.close()
    return jsonify([r['username'] for r in rows])

# Veri değiştiren (POST) isteklerin otomatik denetim kaydı
_LOG_HARIC_YOLLAR = {'/login', '/sifre-degistir'}

@app.after_request
def _islem_logla(response):
    try:
        if request.method == 'POST' and request.path not in _LOG_HARIC_YOLLAR and session.get('user'):
            ozet_parcalari = []
            kaynak = request.get_json(silent=True) or request.form
            if kaynak:
                for k, v in list(kaynak.items())[:8]:
                    deger = str(v)
                    if len(deger) > 60:
                        deger = deger[:60] + '…'
                    ozet_parcalari.append(f"{k}={deger}")
            ozet = ', '.join(ozet_parcalari)[:500]
            db.log_yaz(
                session.get('user'), session.get('role'),
                datetime.now(TR_TZ).strftime('%Y-%m-%d %H:%M:%S'),
                'islem', request.endpoint or '', request.path, ozet, response.status_code
            )
    except Exception:
        pass  # log yazımı asla asıl isteği bozmasın
    return response


# ── Sayfalar ──────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return redirect(url_for('dashboard'))

@app.route('/dashboard')
@login_required
def dashboard():
    return render_template('dashboard.html')

@app.route('/rezervasyonlar')
@login_required
def rezervasyonlar():
    return render_template('rezervasyonlar.html')

@app.route('/rez-formu')
@login_required
def rez_formu():
    return render_template('rez_formu.html')

@app.route('/oda-durumu')
@login_required
def oda_durumu():
    return render_template('oda_durumu.html')

@app.route('/cari')
@login_required
def cari():
    return render_template('cari.html')

@app.route('/adisyon')
@login_required
def adisyon():
    return render_template('adisyon.html')

@app.route('/gunluk-liste')
@login_required
def gunluk_liste():
    return render_template('gunluk_liste.html')

@app.route('/import')
@admin_required
def import_page():
    return render_template('import.html')

@app.route('/yedekleme-talimati')
@admin_required
def yedekleme_talimati():
    return render_template('yedekleme_talimati.html')

@app.route('/api/gunluk-liste')
@login_required
def api_gunluk_liste():
    tarih = request.args.get('tarih', bugun().isoformat())
    conn = db.get_conn()
    # Giriş listesi
    girisler = conn.execute(
        "SELECT oda_no, otel, musteri, yetiskin, cocuk, giris, cikis, foy_no "
        "FROM rezervasyonlar WHERE giris=? AND (durum IS NULL OR durum != 'Kapora Yandı') ORDER BY otel, oda_no", (tarih,)
    ).fetchall()
    # Çıkış listesi — HK kolonları yoksa fallback
    try:
        cikislar = conn.execute(
            "SELECT oda_no, otel, musteri, yetiskin, cocuk, giris, cikis, foy_no, "
            "rez_bakiye, adis_bakiye, anahtar_teslim, anahtar_teslim_zaman, hk_durum "
            "FROM rezervasyonlar WHERE cikis=? AND (durum IS NULL OR durum != 'Kapora Yandı') ORDER BY otel, oda_no", (tarih,)
        ).fetchall()
    except Exception:
        cikislar = conn.execute(
            "SELECT oda_no, otel, musteri, yetiskin, cocuk, giris, cikis, foy_no, "
            "rez_bakiye, adis_bakiye "
            "FROM rezervasyonlar WHERE cikis=? AND (durum IS NULL OR durum != 'Kapora Yandı') ORDER BY otel, oda_no", (tarih,)
        ).fetchall()
    # Kahvaltı listesi: konaklıyorlar (giris < tarih <= cikis) VEYA çıkış günü (cikis = tarih)
    # Giriş günü kahvaltı YOK, çıkış günü VAR
    kahvalti = conn.execute(
        "SELECT oda_no, otel, musteri, yetiskin, cocuk, giris, cikis, foy_no "
        "FROM rezervasyonlar "
        "WHERE giris < ? AND cikis >= ? AND (durum IS NULL OR durum != 'Kapora Yandı') "
        "AND (kahvalti IS NULL OR kahvalti != 'Kahvaltısız') "
        "ORDER BY otel, oda_no", (tarih, tarih)
    ).fetchall()
    conn.close()
    def row2dict(r, include_hk=False):
        d = {
            'oda_no': r['oda_no'], 'otel': r['otel'], 'musteri': r['musteri'],
            'yetiskin': r['yetiskin'], 'cocuk': r['cocuk'],
            'giris': r['giris'], 'cikis': r['cikis'], 'foy_no': r['foy_no']
        }
        if include_hk:
            def _safe_num(key):
                try: v = r[key]; return float(v) if v else 0
                except Exception: return 0
            def _safe_str(key):
                try: v = r[key]; return str(v) if v else ''
                except Exception: return ''
            def _safe_int(key):
                try: v = r[key]; return int(v) if v else 0
                except Exception: return 0
            d['rez_bakiye']           = _safe_num('rez_bakiye')
            d['adis_bakiye']          = _safe_num('adis_bakiye')
            d['toplam_bakiye']        = _safe_num('rez_bakiye') + _safe_num('adis_bakiye')
            d['anahtar_teslim']       = _safe_int('anahtar_teslim')
            d['anahtar_teslim_zaman'] = _safe_str('anahtar_teslim_zaman')
            d['hk_durum']             = _safe_str('hk_durum')
        return d
    return jsonify({
        'tarih': tarih,
        'girisler': [row2dict(r) for r in girisler],
        'cikislar': [row2dict(r, include_hk=True) for r in cikislar],
        'kahvalti': [row2dict(r) for r in kahvalti]
    })


# ── Anahtar Teslim & Housekeeping ────────────────────────────────────────────

@app.route('/api/anahtar-teslim', methods=['POST'])
@login_required
def api_anahtar_teslim():
    data   = request.get_json()
    foy_no = data.get('foy_no')
    isle   = data.get('isle', True)
    force  = data.get('force', False)
    conn = db.get_conn()
    rez = conn.execute(
        "SELECT foy_no, rez_bakiye, adis_bakiye, oda_no, otel, musteri FROM rezervasyonlar WHERE foy_no=?",
        (foy_no,)
    ).fetchone()
    if not rez:
        conn.close()
        return jsonify({'ok': False, 'error': 'Rezervasyon bulunamadı'})
    toplam_bakiye = (rez['rez_bakiye'] or 0) + (rez['adis_bakiye'] or 0)
    if isle and toplam_bakiye > 0 and not force:
        conn.close()
        return jsonify({'ok': False, 'bakiye_uyari': True, 'bakiye': toplam_bakiye,
                        'oda_no': rez['oda_no'], 'musteri': rez['musteri']})
    from datetime import datetime
    from zoneinfo import ZoneInfo
    now_str = datetime.now(ZoneInfo('Europe/Istanbul')).strftime('%Y-%m-%d %H:%M')
    if isle:
        conn.execute(
            "UPDATE rezervasyonlar SET anahtar_teslim=1, anahtar_teslim_zaman=?, hk_durum='temizleniyor' WHERE foy_no=?",
            (now_str, foy_no)
        )
        try:
            kullanici = session.get('username', '?')
            conn.execute(
                "INSERT INTO islem_loglari(kullanici, islem_turu, tablo, kayit_id, aciklama) VALUES(?,?,?,?,?)",
                (kullanici, 'anahtar_teslim', 'rezervasyonlar', foy_no,
                 f"Föy#{foy_no} Oda {rez['oda_no']} - Anahtar teslim{(' (BAKİYELİ: ₺'+str(toplam_bakiye)+')') if toplam_bakiye > 0 else ''}")
            )
        except Exception:
            pass
        conn.commit()
        conn.close()
        # Telegram bildirimi
        mesaj = f"{rez['otel']} - Oda {rez['oda_no']}\nOda Boşaldı. Temizlik yapılıp bilgi verilecek!"
        telegram_gonder(mesaj)
    else:
        conn.execute(
            "UPDATE rezervasyonlar SET anahtar_teslim=0, anahtar_teslim_zaman='', hk_durum='' WHERE foy_no=?",
            (foy_no,)
        )
        conn.commit()
        conn.close()
    return jsonify({'ok': True, 'isle': isle})


@app.route('/hk-listesi')
@login_required
def hk_listesi():
    return render_template('hk_listesi.html')


@app.route('/api/hk-listesi')
@login_required
def api_hk_listesi():
    tarih = request.args.get('tarih', bugun().isoformat())
    conn  = db.get_conn()
    try:
        rows  = conn.execute(
            "SELECT foy_no, oda_no, otel, musteri, cikis, anahtar_teslim_zaman, hk_durum, "
            "rez_bakiye, adis_bakiye "
            "FROM rezervasyonlar "
            "WHERE cikis=? AND anahtar_teslim=1 "
            "ORDER BY hk_durum DESC, otel, oda_no",
            (tarih,)
        ).fetchall()
    except Exception:
        rows = []
    conn.close()
    return jsonify({'tarih': tarih, 'odalar': [
        {
            'foy_no':   r['foy_no'],
            'oda_no':   r['oda_no'],
            'otel':     r['otel'],
            'musteri':  r['musteri'],
            'cikis':    r['cikis'],
            'teslim_zaman': r['anahtar_teslim_zaman'] or '',
            'hk_durum': r['hk_durum'] or 'bekliyor',
            'bakiyeli': ((r['rez_bakiye'] or 0) + (r['adis_bakiye'] or 0)) > 0,
        } for r in rows
    ]})


import os
TELEGRAM_TOKEN   = os.environ.get('TELEGRAM_TOKEN', '') or cfg.get('sistem.telegram_token', '')
TELEGRAM_CHAT_ID = os.environ.get('TELEGRAM_CHAT_ID', '') or cfg.get('sistem.telegram_chat_id', '')

def telegram_gonder(mesaj):
    import threading
    def _gonder():
        try:
            import requests as _req
            url = f'https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage'
            r = _req.post(url, json={'chat_id': TELEGRAM_CHAT_ID, 'text': mesaj}, timeout=10)
            print(f'Telegram yanit: {r.status_code} {r.text[:100]}')
        except Exception as e:
            print(f'Telegram hata: {e}')
    threading.Thread(target=_gonder, daemon=True).start()


@app.route('/telegram-webhook', methods=['POST'])
def telegram_webhook():
    """Telegram'dan gelen komutları işler: /temiz 18, /temizleniyor 18, /bekliyor 18"""
    try:
        data = request.get_json(force=True)
        msg  = data.get('message', {})
        text = msg.get('text', '').strip()
        chat_id = str(msg.get('chat', {}).get('id', ''))

        # Sadece HK grubundan gelen komutları kabul et
        if chat_id != TELEGRAM_CHAT_ID:
            return jsonify({'ok': True})

        # Komut parse:
        # Kısa format: 18T=temiz, 18B=başladı(temizleniyor), 18G=geri al(bekliyor)
        # Uzun format:  /temiz 18, /temizleniyor 18, /bekliyor 18
        import re
        txt = text.strip().lower()
        m_kisa = re.match(r'^(\d+)(t|b|g)$', txt)
        m_uzun = re.match(r'^/(temiz|temizleniyor|bekliyor)\s+(\d+)$', txt)

        if m_kisa:
            oda_no = int(m_kisa.group(1))
            harf   = m_kisa.group(2)
            komut  = {'t': 'temiz', 'b': 'temizleniyor', 'g': 'bekliyor'}[harf]
        elif m_uzun:
            komut  = m_uzun.group(1)
            oda_no = int(m_uzun.group(2))
        else:
            return jsonify({'ok': True})

        conn = db.get_conn()
        rez  = conn.execute(
            "SELECT foy_no, oda_no, otel, cikis FROM rezervasyonlar "
            "WHERE oda_no=? AND anahtar_teslim=1 AND hk_durum != 'temiz' "
            "ORDER BY cikis DESC LIMIT 1",
            (oda_no,)
        ).fetchone()

        if not rez:
            conn.close()
            telegram_gonder(f"⚠️ Oda {oda_no} için aktif HK kaydı bulunamadı.")
            return jsonify({'ok': True})

        conn.execute(
            "UPDATE rezervasyonlar SET hk_durum=? WHERE foy_no=?",
            (komut, rez['foy_no'])
        )
        conn.commit()
        conn.close()

        durum_tr = {'temiz': '✅ Temiz', 'temizleniyor': '🔄 Temizleniyor', 'bekliyor': '⏳ Bekliyor'}
        from datetime import datetime
        from zoneinfo import ZoneInfo
        simdi = datetime.now(ZoneInfo('Europe/Istanbul')).strftime('%H:%M')
        onay = f"{durum_tr[komut]} — {rez['otel']} Oda {oda_no} güncellendi. ⏰ {simdi}"
        telegram_gonder(onay)

    except Exception as e:
        print(f'Webhook hata: {e}')

    return jsonify({'ok': True})


@app.route('/api/hk-durum', methods=['POST'])
@login_required
def api_hk_durum():
    data   = request.get_json()
    foy_no = data.get('foy_no')
    durum  = data.get('durum')
    if durum not in ('temizleniyor', 'temiz', 'bekliyor'):
        return jsonify({'ok': False, 'error': 'Geçersiz durum'})
    conn = db.get_conn()
    rez  = conn.execute(
        "SELECT oda_no, otel FROM rezervasyonlar WHERE foy_no=?", (foy_no,)
    ).fetchone()
    conn.execute("UPDATE rezervasyonlar SET hk_durum=? WHERE foy_no=?", (durum, foy_no))
    conn.commit()
    conn.close()

    # Telegram bildirimi — sadece Temizlemeye Başla ve Temizlendi'de
    if rez and durum in ('temizleniyor', 'temiz'):
        from datetime import datetime
        from zoneinfo import ZoneInfo
        simdi = datetime.now(ZoneInfo('Europe/Istanbul')).strftime('%H:%M')
        if durum == 'temizleniyor':
            mesaj = f"{rez['otel']} - Oda {rez['oda_no']}\nOda Boşaldı. Temizlik yapılıp bilgi verilecek!"
        else:
            mesaj = f"✅ {rez['otel']} - Oda {rez['oda_no']}\nTemizlik Tamamlandı. ⏰ {simdi}"
        telegram_gonder(mesaj)

    return jsonify({'ok': True})


# ── API — Okuma ───────────────────────────────────────────────────────────────

@app.route('/api/dashboard')
def api_dashboard():
    today = bugun().isoformat()
    girisler, cikislar, aktifler, kahvalti = db.get_dashboard(today)
    kisa_ad   = cfg.otel_bilgi().get('kisa_ad', 'OTEL')
    leo_aktif = sum(1 for r in aktifler if r['otel'] == kisa_ad)
    cv_aktif  = 0
    toplam_alacak = sum((r.get('rez_bakiye') or 0) + (r.get('adis_bakiye') or 0)
                        for r in db.get_rezervasyonlar()
                        if r.get('durum') != 'Kapora Yandı')
    return jsonify({
        'today': tr_tarih(bugun()),
        'stats': {
            'bugun_giris':   len(girisler),
            'bugun_cikis':   len(cikislar),
            'leo_aktif':     leo_aktif,
            'cv_aktif':      cv_aktif,
            'kahvalti_kisi': kahvalti,
            'toplam_alacak': toplam_alacak,
            'toplam_oda':    cfg.otel_bilgi().get('toplam_oda', 20),
            'otel_adi':      cfg.otel_bilgi().get('ad', 'Otel'),
        },
        'girisler': girisler,
        'cikislar': cikislar,
        'aktifler': aktifler,
    })

@app.route('/api/rezervasyonlar')
def api_rezervasyonlar():
    q     = request.args.get('q', '')
    otel  = request.args.get('otel', 'Tümü')
    return jsonify(db.get_rezervasyonlar(q=q, otel=otel))

@app.route('/api/oda-durumu')
def api_oda_durumu():
    from datetime import timedelta
    start_str = request.args.get('start', bugun().isoformat())
    try:
        start = date.fromisoformat(start_str)
    except:
        start = bugun()
    days  = 30
    dates = [start + timedelta(i) for i in range(days)]
    today = bugun().isoformat()

    rezervasyonlar = [r for r in db.get_rezervasyonlar() if r.get("durum") != "Kapora Yandı"]
    otel_cfg   = cfg.otel_bilgi()
    bas        = otel_cfg.get('oda_baslangic', 1)
    bit        = otel_cfg.get('oda_bitis', 20) + 1
    kisa_ad    = otel_cfg.get('kisa_ad', 'OTEL')
    all_rooms  = [(kisa_ad, o) for o in range(bas, bit)]

    grid = []
    for otel_label, oda_no in all_rooms:
        cells = []
        for d in dates:
            ds = d.isoformat()
            rez = next((r for r in rezervasyonlar
                        if r['oda_no'] == oda_no
                        and r['giris'] and r['cikis']
                        and r['giris'] <= ds < r['cikis']), None)
            if rez:
                parts = rez['musteri'].split()
                initials = ' '.join(p[0] for p in parts[:2]) if parts else '?'
                cells.append({
                    'dolu': True, 'initials': initials,
                    'musteri': rez['musteri'],
                    'giris': rez['giris'], 'cikis': rez['cikis'],
                    'rez_bakiye': rez.get('rez_bakiye') or 0,
                    'foy_no': rez['foy_no'],
                    'is_giris': ds == rez['giris'],
                    'otel': rez['otel'],
                })
            else:
                cells.append({'dolu': False})
        grid.append({'otel': otel_label, 'oda_no': oda_no, 'cells': cells})

    return jsonify({'dates': [d.isoformat() for d in dates], 'today': today, 'grid': grid})

@app.route('/api/musaitlik')
def api_musaitlik():
    # Verilen giriş/çıkış aralığında ve (varsa) otelde boş odaları döner. Salt-okunur, mevcut hiçbir kayda dokunmaz.
    giris = request.args.get('giris', '')
    cikis = request.args.get('cikis', '')
    otel  = request.args.get('otel', 'Tümü')
    if not giris or not cikis or cikis <= giris:
        return jsonify({'odalar': [], 'error': 'Geçersiz tarih aralığı'})

    rezervasyonlar = [r for r in db.get_rezervasyonlar() if r.get('durum') != 'Kapora Yandı']
    otel_cfg   = cfg.otel_bilgi()
    bas        = otel_cfg.get('oda_baslangic', 1)
    bit        = otel_cfg.get('oda_bitis', 20) + 1
    kisa_ad    = otel_cfg.get('kisa_ad', 'OTEL')
    leo_odalar = [(kisa_ad, o) for o in range(bas, bit)]
    cv_odalar  = []
    all_rooms  = cv_odalar + leo_odalar
    if otel in ([cfg.otel_bilgi().get('kisa_ad', 'OTEL')]):
        all_rooms = [r for r in all_rooms if r[0] == otel]

    musait = []
    for otel_label, oda_no in all_rooms:
        cakisan = next((r for r in rezervasyonlar
                         if r['oda_no'] == oda_no
                         and r['giris'] and r['cikis']
                         and r['giris'] < cikis and r['cikis'] > giris), None)
        if not cakisan:
            musait.append({'otel': otel_label, 'oda_no': oda_no})
    return jsonify({'odalar': musait})

@app.route('/api/cari')
def api_cari():
    rezervasyonlar = db.get_rezervasyonlar()
    def _f(v): return float(v or 0)
    aktif_rez = [r for r in rezervasyonlar if r.get('durum') != 'Kapora Yandı']
    def _tahsilat(r):
        return _f(r.get('rez_tahsilat')) + _f(r.get('kapora'))

    # Föy başına adisyon kayıt sayısı (İşlem Detayı rozeti için)
    tum_adisyonlar = db.get_adisyonlar()
    adis_sayilari = {}
    for a in tum_adisyonlar:
        fn = a.get('foy_no')
        adis_sayilari[fn] = adis_sayilari.get(fn, 0) + 1
    for r in rezervasyonlar:
        r['adisyon_sayisi'] = adis_sayilari.get(r.get('foy_no'), 0)

    ozet = {
        'toplam_rez':      sum(_f(r.get('toplam_fiyat')) for r in aktif_rez),
        'toplam_tahsilat': sum(_tahsilat(r) for r in aktif_rez),
        'rez_bakiye':      sum(max(0, _f(r.get('toplam_fiyat')) - _tahsilat(r)) for r in aktif_rez),
        'adis_toplam':     sum(_f(r.get('adisyon')) for r in aktif_rez),
        'adis_tahsilat':   sum(_f(r.get('adis_tahsilat')) for r in aktif_rez),
        'adis_bakiye':     sum(_f(r.get('adis_bakiye')) for r in aktif_rez),
    }
    return jsonify({'ozet': ozet, 'rezervasyonlar': rezervasyonlar})

@app.route('/api/adisyonlar')
def api_adisyonlar():
    foy_f = request.args.get('foy')
    adisyonlar = db.get_adisyonlar(foy_no=int(foy_f) if foy_f else None)
    rezervasyonlar = db.get_rezervasyonlar()
    musteri_map = {r['foy_no']: r['musteri'] for r in rezervasyonlar}
    for a in adisyonlar:
        a['musteri'] = musteri_map.get(a['foy_no'], '')
    aktif_rez = [r for r in rezervasyonlar if r.get('durum') != 'Kapora Yandı']
    foy_listesi = [{'foy_no': r['foy_no'], 'musteri': r['musteri'], 'oda_no': r['oda_no']}
                   for r in sorted(aktif_rez, key=lambda x: x['foy_no'])]
    return jsonify({'adisyonlar': adisyonlar, 'foy_listesi': foy_listesi})

@app.route('/api/next-foy')
def api_next_foy():
    return jsonify({'foy_no': db.get_next_foy_no()})

@app.route('/api/next-adisyon')
def api_next_adisyon():
    return jsonify({'adisyon_no': db.get_next_adisyon_no()})


# ── API — Yazma ───────────────────────────────────────────────────────────────

@app.route('/api/rezervasyon/yeni', methods=['POST'])
def api_rez_yeni():
    try:
        data = request.get_json()
        db.save_rezervasyon(data)
        # Yevmiye: konaklama geliri
        foy_no = int(data.get('foy_no') or 0)
        yevmiye_rez_kaydet(
            foy_no, float(data.get('toplam_fiyat') or 0),
            data.get('otel', cfg.otel_bilgi().get('kisa_ad','OTEL')), data.get('giris'),
            data.get('musteri', ''),
            kapora=float(data.get('kapora') or 0),
            kapora_tarihi=data.get('kapora_tarihi')
        )
        acente_kod = KANAL_MAP.get(data.get('kanal', ''), data.get('kanal', ''))
        if acente_kod in ACENTE_OTO_KODLAR:
            acente_oto_kaydet(foy_no, float(data.get('toplam_fiyat') or 0),
                              data.get('otel', cfg.otel_bilgi().get('kisa_ad','OTEL')),
                              data.get('giris') or bugun().isoformat(),
                              data.get('musteri', ''), acente_kod)
        else:
            acente_oto_temizle(foy_no)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

@app.route('/api/rezervasyon/sil', methods=['POST'])
def api_rez_sil():
    try:
        d = request.get_json()
        foy_no = int(d['foy_no'])
        # Yevmiyeden ilgili kayıtları sil
        conn = mdb.get_conn()
        conn.execute("DELETE FROM yevmiye WHERE aciklama LIKE ?", (f'Föy#{foy_no} %',))
        conn.commit(); conn.close()
        acente_cari_oto_sil(foy_no)
        # Rezervasyonu sil
        db.delete_rezervasyon(foy_no)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

@app.route('/api/rezervasyon/checkin', methods=['POST'])
@login_required
def api_checkin():
    try:
        d = request.get_json()
        foy_no = int(d['foy_no'])
        rez = db.get_rezervasyonlar()
        r = next((x for x in rez if x['foy_no'] == foy_no), None)
        if not r:
            return jsonify({'ok': False, 'error': 'Rezervasyon bulunamadı'}), 404
        if r.get('checkin'):
            return jsonify({'ok': False, 'error': 'Check-in zaten yapılmış'}), 400

        otel = r.get('otel', cfg.otel_bilgi().get('kisa_ad','OTEL'))
        musteri = r.get('musteri', '')
        toplam = float(r.get('toplam_fiyat') or 0)
        kapora = float(r.get('kapora') or 0)
        tarih = bugun().isoformat()
        gelir_hesap = '600'
        aciklama = f'Föy#{foy_no} {musteri} check-in'

        # Check-in kaydını güncelle
        conn_db = db.get_conn()
        conn_db.execute("UPDATE rezervasyonlar SET checkin=1, durum='Konaklıyor' WHERE foy_no=?", (foy_no,))
        conn_db.commit(); conn_db.close()

        # Yevmiye:
        # A. Gelir doğumu: 120 borç / 601 alacak (toplam konaklama)
        # B. Kapora mahsubu: 340 borç / 120 alacak
        conn = mdb.get_conn()
        if toplam > 0:
            mdb._yevmiye_ekle(conn, tarih, 'Check-in Konaklama', '120', gelir_hesap,
                              toplam, aciklama, otel)
        # Kapora mahsubu kaldırıldı - kapora zaten 102-1/120 ile yazıldı
        conn.commit(); conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

@app.route('/api/rezervasyon/checkin-iptal', methods=['POST'])
@login_required
def api_checkin_iptal():
    try:
        d = request.get_json()
        foy_no = int(d['foy_no'])
        rez = db.get_rezervasyonlar()
        r = next((x for x in rez if x['foy_no'] == foy_no), None)
        if not r or not r.get('checkin'):
            return jsonify({'ok': False, 'error': 'Check-in bulunamadı'}), 400

        otel = r.get('otel', cfg.otel_bilgi().get('kisa_ad','OTEL'))
        musteri = r.get('musteri', '')
        toplam = float(r.get('toplam_fiyat') or 0)
        kapora = float(r.get('kapora') or 0)
        tarih = bugun().isoformat()
        gelir_hesap = '600'
        aciklama = f'Föy#{foy_no} {musteri} check-in iptal (storno)'

        # Check-in iptal
        conn_db = db.get_conn()
        conn_db.execute("UPDATE rezervasyonlar SET checkin=0, durum='Aktif' WHERE foy_no=?", (foy_no,))
        conn_db.commit(); conn_db.close()

        # Storno kayıtları (ters çevir)
        conn = mdb.get_conn()
        if toplam > 0:
            mdb._yevmiye_ekle(conn, tarih, 'Check-in Storno', gelir_hesap, '120',
                              toplam, aciklama, otel)
        if kapora > 0:
            mdb._yevmiye_ekle(conn, tarih, 'Kapora Mahsubu Storno', '120', '102-1',
                              kapora, aciklama, otel)
        conn.commit(); conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

@app.route('/api/rezervasyon/kapora-yandi', methods=['POST'])
def api_kapora_yandi():
    try:
        d = request.get_json()
        foy_no = int(d['foy_no'])
        rezler = db.get_rezervasyonlar()
        r = next((x for x in rezler if x['foy_no'] == foy_no), None)
        if not r:
            return jsonify({'ok': False, 'error': 'Rezervasyon bulunamadı'}), 404
        kapora = float(r.get('kapora') or 0)
        if kapora <= 0:
            return jsonify({'ok': False, 'error': 'Kapora yok'}), 400
        otel = r.get('otel', cfg.otel_bilgi().get('kisa_ad','OTEL'))
        gelir_hesap = '600'
        musteri = r.get('musteri', '')
        tarih = bugun().isoformat()
        # Yevmiye: Alınan Avanslar borç / Diğer Olağan Gelir alacak (340/649)
        conn = mdb.get_conn()
        mdb._yevmiye_ekle(conn, tarih, 'Kapora Yanması', '120', '649',
                          kapora, f'Föy#{foy_no} {musteri} kapora yandı - iptal bedeli', otel)
        conn.commit(); conn.close()
        # Rezervasyon durumunu güncelle, bakiyeyi sıfırla (silinmez)
        otel_conn = db.get_conn()
        otel_conn.execute(
            "UPDATE rezervasyonlar SET durum='Kapora Yandı', rez_bakiye=0 WHERE foy_no=?",
            (foy_no,)
        )
        otel_conn.commit(); otel_conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400


@app.route('/api/rezervasyon/guncelle', methods=['POST'])
def api_rez_guncelle():
    try:
        data = request.get_json()
        db.update_rezervasyon(data['foy_no'], data)
        # Yevmiye: konaklama geliri güncelle
        foy_no = int(data.get('foy_no') or 0)
        yevmiye_rez_kaydet(
            foy_no, float(data.get('toplam_fiyat') or 0),
            data.get('otel', cfg.otel_bilgi().get('kisa_ad','OTEL')), data.get('giris'),
            data.get('musteri', ''),
            kapora=float(data.get('kapora') or 0),
            kapora_tarihi=data.get('kapora_tarihi'),
            guncelleme=True
        )
        acente_kod = KANAL_MAP.get(data.get('kanal', ''), data.get('kanal', ''))
        if acente_kod in ACENTE_OTO_KODLAR:
            acente_oto_kaydet(foy_no, float(data.get('toplam_fiyat') or 0),
                              data.get('otel', cfg.otel_bilgi().get('kisa_ad','OTEL')),
                              data.get('giris') or bugun().isoformat(),
                              data.get('musteri', ''), acente_kod)
        else:
            acente_oto_temizle(foy_no)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400


def yevmiye_rez_kaydet(foy_no, toplam_fiyat, otel, giris, musteri,
                       kapora=0, kapora_tarihi=None, guncelleme=False):
    """Rezervasyon konaklama geliri ve kaporayı yevmiyeye yazar."""
    try:
        tarih = giris or bugun().isoformat()
        gelir_hesap = '600'
        aciklama_kon = f'Föy#{foy_no} {musteri} konaklama'
        aciklama_kap = f'Föy#{foy_no} {musteri} kapora'
        conn = mdb.get_conn()
        # Eski kayıtları sil - foy_no ile eşleştir (misafir adı değişse bile çalışır)
        conn.execute("DELETE FROM yevmiye WHERE aciklama LIKE ? AND islem_tipi IN ('Konaklama Geliri','Kapora')",
                     (f'Föy#{foy_no} %',))
        conn.commit()

        kap_tarih = kapora_tarihi or tarih

        # Konaklama geliri: Müşteri Cari borç / Konaklama Geliri alacak
        if toplam_fiyat and toplam_fiyat > 0:
            mdb._yevmiye_ekle(conn, tarih, 'Konaklama Geliri', '120', gelir_hesap,
                              toplam_fiyat, aciklama_kon, otel)

        # Kapora: İş Bankası borç / Müşteri Cari alacak (120)
        if kapora and kapora > 0:
            mdb._yevmiye_ekle(conn, kap_tarih, 'Kapora', '102-1', '120',
                              kapora, aciklama_kap, otel)

        conn.commit(); conn.close()
    except Exception as e:
        print(f'Yevmiye rez kayıt hatası: {e}')


def acente_cari_oto_sil(foy_no):
    """Rezervasyon silinince acente_cari kaydını da siler."""
    try:
        conn = mdb.get_conn()
        conn.execute("DELETE FROM acente_cari WHERE foy_no=?", (foy_no,))
        conn.commit(); conn.close()
    except Exception as e:
        print(f'Acente oto sil hatası: {e}')


def acente_oto_kaydet(foy_no, toplam_fiyat, otel, tarih, musteri, acente_kod):
    """Acente üzerinden gelen (BKG/JLY/TTS/ETS) rezervasyon kaydedilince/güncellenince
    otomatik olarak:
      1) {acente_hesap} Borç / 120 Müşteri Cari Alacak  (rez. bedeli — tam tutar)
      2) 730 Acente Komisyonu Borç / {acente_hesap} Alacak (komisyon — acenteye net
         fatura kesildiği için bedel hemen düşülür)
    Müşterinin borcu acenteye devredilmiş olur; ayrıca elle 'Ödeme Al' yapılmasına
    gerek kalmaz. Föy'e ait önceki otomatik kayıtlar silinip yeniden yazılır (idempotent)."""
    hesap = ACENTE_HESAP_KODU.get(acente_kod)
    adi = ACENTE_ADI.get(acente_kod, acente_kod)
    if not hesap:
        return
    try:
        conn = mdb.get_conn()
        # Önceki otomatik kayıtları temizle
        conn.execute("DELETE FROM yevmiye WHERE aciklama LIKE ? AND aciklama LIKE '%[ACENTE-OTO]%'",
                     (f'Föy#{foy_no} %',))

        if toplam_fiyat and toplam_fiyat > 0:
            aciklama = f'Föy#{foy_no} {musteri} [ACENTE-OTO] rez tahsilat'
            mdb._yevmiye_ekle(conn, tarih, f'Acente Tahsilat ({adi})',
                              hesap, '120', toplam_fiyat, aciklama, otel)

            a = conn.execute("SELECT komisyon_orani FROM acenteler WHERE kod=?", (acente_kod,)).fetchone()
            oran = float(a['komisyon_orani']) if a else 15.0
            komisyon = round(toplam_fiyat * oran / 100, 2)
            if komisyon > 0:
                aciklama_kom = f'Föy#{foy_no} {musteri} [ACENTE-OTO] komisyon'
                mdb._yevmiye_ekle(conn, tarih, f'Acente Komisyonu ({adi})',
                                  '730', hesap, komisyon, aciklama_kom, otel)

        conn.commit(); conn.close()

        # Rezervasyon tahsilat alanlarını da senkronize et — desk'te ayrıca
        # ödeme alınmasına gerek kalmasın
        otel_conn = db.get_conn()
        otel_conn.execute("""
            UPDATE rezervasyonlar SET rez_tahsilat=?, rez_bakiye=0, rez_odeme_sekli=?
            WHERE foy_no=?
        """, (toplam_fiyat or 0, f'Acente ({adi})', foy_no))
        otel_conn.commit(); otel_conn.close()
    except Exception as e:
        print(f'Acente oto kayıt hatası ({acente_kod}): {e}')


def acente_oto_temizle(foy_no):
    """Bir rezervasyonun acentesi otomatik-muhasebe kapsamından çıktığında
    (veya föy silindiğinde) otomatik acente kayıtlarını temizler ve rezervasyonun
    tahsilat alanlarını sıfırlar (yalnızca tahsilat 'Acente (...)' ise — elle
    girilmiş gerçek bir tahsilat varsa dokunulmaz)."""
    try:
        conn = mdb.get_conn()
        conn.execute("DELETE FROM yevmiye WHERE aciklama LIKE ? AND aciklama LIKE '%[ACENTE-OTO]%'",
                     (f'Föy#{foy_no} %',))
        conn.commit(); conn.close()

        otel_conn = db.get_conn()
        r = otel_conn.execute(
            "SELECT toplam_fiyat, rez_odeme_sekli FROM rezervasyonlar WHERE foy_no=?", (foy_no,)
        ).fetchone()
        if r and (r['rez_odeme_sekli'] or '').startswith('Acente ('):
            otel_conn.execute("""
                UPDATE rezervasyonlar SET rez_tahsilat=0, rez_bakiye=?, rez_odeme_sekli=''
                WHERE foy_no=?
            """, (r['toplam_fiyat'] or 0, foy_no))
            otel_conn.commit()
        otel_conn.close()
    except Exception as e:
        print(f'Acente oto temizleme hatası: {e}')


def migrate_acente_otomatik():
    """Tek seferlik migration: var olan tüm acente (BKG/JLY/TTS/ETS) rezervasyonları
    için eski/elle girilmiş tahsilat-komisyon kayıtlarını temizleyip yeni otomatik
    akışı uygular. parametreler tablosundaki bayrak sayesinde bir daha çalışmaz;
    gerekirse o satır silinip yeniden tetiklenebilir."""
    FLAG = 'migration_acente_oto_v2'
    try:
        mconn = mdb.get_conn()
        if mconn.execute("SELECT 1 FROM parametreler WHERE anahtar=?", (FLAG,)).fetchone():
            mconn.close()
            return
        rconn = db.get_conn()
        rezler = rconn.execute(
            "SELECT foy_no, kanal, toplam_fiyat, otel, giris, musteri FROM rezervasyonlar"
        ).fetchall()
        rconn.close()
        count = 0
        for r in rezler:
            acente_kod = KANAL_MAP.get(r['kanal'] or '', r['kanal'] or '')
            if acente_kod not in ACENTE_OTO_KODLAR:
                continue
            foy_no = r['foy_no']
            # Bu föy'e ait eski (manuel/yanlış) tahsilat & komisyon kayıtlarını temizle
            mconn.execute("""
                DELETE FROM yevmiye
                WHERE aciklama LIKE ?
                  AND (islem_tipi LIKE '%Tahsilat%' OR islem_tipi LIKE '%Komisyon%')
                  AND aciklama NOT LIKE '%[ACENTE-OTO]%'
                  AND aciklama NOT LIKE '%[ACENTE-FATURA]%'
            """, (f'Föy#{foy_no} %',))
            mconn.execute("DELETE FROM acente_cari WHERE foy_no=?", (foy_no,))
            mconn.commit()
            acente_oto_kaydet(
                foy_no, float(r['toplam_fiyat'] or 0), r['otel'] or cfg.otel_bilgi().get('kisa_ad','OTEL'),
                r['giris'] or bugun().isoformat(), r['musteri'] or '', acente_kod
            )
            count += 1
        mconn.execute("INSERT OR REPLACE INTO parametreler(anahtar,deger) VALUES(?,?)",
                      (FLAG, f'{count} föy güncellendi — {datetime.now().isoformat()}'))
        mconn.commit(); mconn.close()
        print(f'[Migration] Acente otomatik muhasebe (BKG/JLY/TTS/ETS): {count} rezervasyon güncellendi.')
    except Exception as e:
        print(f'[Migration] Acente oto hata: {e}')

ODEME_HESAP_KODU = {
    'Nakit': '100', 'nakit': '100',
    'Kredi Kartı': '102-1', 'KK': '102-1', 'kk': '102-1',
    'İş Bankası': '102-1', 'Denizbank': '102-3',
}



@app.route('/api/tahsilat/gecmis')
def api_tahsilat_gecmis():
    foy_no = request.args.get('foy_no', type=int)
    conn = mdb.get_conn()
    rows = conn.execute("""
        SELECT id, tarih, islem_tipi, tutar, borc_hesap
        FROM yevmiye 
        WHERE (islem_tipi LIKE '%Tahsilat%' OR islem_tipi = 'Kapora' OR islem_tipi = 'Kapora Yanması') 
        AND aciklama LIKE ?
        ORDER BY tarih ASC
    """, (f'Föy#{foy_no} %',)).fetchall()
    conn.close()
    odeme_map = {
        '100': 'Nakit', '102-1': 'İş Bankası', '102-2': 'Ziraat',
        '102-3': 'Denizbank', '120': 'Müşteri Cari', '340': 'Alınan Kaparo',
        '320-1': 'Booking', '320-2': 'Expedia',
        '320-3': 'JollyTur', '320-4': 'TatilSepeti', '320-5': 'ETSTUR',
    }
    result = []
    for r in rows:
        islem_tipi = r[2]
        # islem_tipi'nden ödeme adını çıkar: 'Rezervasyon Tahsilat - Kredi Kartı' → 'Kredi Kartı'
        if ' - ' in islem_tipi:
            odeme_adi = islem_tipi.split(' - ', 1)[1]
        elif islem_tipi == 'Kapora':
            odeme_adi = 'İş Bankası'
        else:
            odeme_adi = odeme_map.get(r[4], r[4])
        result.append({
            'id': r[0], 'tarih': r[1], 'tur': r[2],
            'tutar': r[3], 'odeme': odeme_adi,
            'foy_no': foy_no
        })
    return jsonify(result)

@app.route('/api/tahsilat/sil', methods=['POST'])
def api_tahsilat_sil():
    try:
        d = request.get_json()
        yev_id = int(d.get('yev_id') or d.get('id', 0))
        foy_no = int(d['foy_no']) if d.get('foy_no') else None
        tutar  = float(d.get('tutar', 0))
        tur    = d.get('tur', '')

        # Yevmiyeyi sil
        conn = mdb.get_conn()
        conn.execute("DELETE FROM yevmiye WHERE id=?", (yev_id,))
        conn.commit(); conn.close()

        # Rezervasyon tablosunu güncelle
        if foy_no and tutar > 0:
            otel_conn = db.get_conn()
            if 'Adisyon' in tur:
                # adisyon_odemeler tablosundan da sil
                # yevmiye aciklamasından adisyon_no'yu bul
                mconn = mdb.get_conn()
                row = mconn.execute(
                    "SELECT aciklama FROM yevmiye WHERE id=?", (yev_id,)
                ).fetchone()
                mconn.close()
                # adisyon_odemeler'de foy_no + tutar + eşleşen kaydı sil
                otel_conn.execute("""
                    DELETE FROM adisyon_odemeler
                    WHERE foy_no=? AND tutar=?
                    AND id=(SELECT id FROM adisyon_odemeler WHERE foy_no=? AND tutar=? LIMIT 1)
                """, (foy_no, tutar, foy_no, tutar))
                otel_conn.execute("""
                    UPDATE rezervasyonlar
                    SET adis_tahsilat = MAX(0, adis_tahsilat - ?),
                        adis_bakiye   = adis_bakiye + ?
                    WHERE foy_no=?
                """, (tutar, tutar, foy_no))
                # adisyon toplamlarını senkronize et
                db._sync_adisyon_totals(otel_conn, foy_no)
            else:
                otel_conn.execute("""
                    UPDATE rezervasyonlar
                    SET rez_tahsilat = MAX(0, rez_tahsilat - ?),
                        rez_bakiye   = rez_bakiye + ?
                    WHERE foy_no=?
                """, (tutar, tutar, foy_no))
            otel_conn.commit(); otel_conn.close()

        # Sync sonrası yeniden hesapla
        if foy_no:
            sc = db.get_conn()
            db.sync_rez_tahsilat(sc, foy_no)
            db._sync_adisyon_totals(sc, foy_no)
            sc.commit(); sc.close()

        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400
@app.route('/api/tahsilat/rez', methods=['POST'])
def api_rez_tah():
    try:
        d = request.get_json()
        foy_no = int(d['foy_no'])
        tutar = float(d['tutar'])
        odeme = d['odeme']
        db.save_rez_tahsilat(foy_no, tutar, odeme)
        # Yevmiye: otomatik kayıt
        hesap_kodu = ODEME_HESAP_KODU.get(odeme)
        if hesap_kodu and tutar > 0:
            rez = db.get_rezervasyonlar()
            r = next((x for x in rez if x['foy_no'] == foy_no), None)
            tarih = d.get('tarih') or bugun().isoformat()
            otel = r.get('otel', cfg.otel_bilgi().get('kisa_ad','OTEL')) if r else 'LEO'
            musteri = r.get('musteri', '') if r else ''
            gelir_hesap = '600'
            conn = mdb.get_conn()
            mdb._yevmiye_ekle(conn, tarih, 'Rezervasyon Tahsilat - ' + odeme,
                              hesap_kodu, '120', tutar,
                              f'Föy#{foy_no} {musteri} rez tahsilat', otel)
            # Kapora mahsubu check-in anında yazılıyor, tahsilatta yazılmaz
            conn.commit(); conn.close()
        # Sync: rez_tahsilat yeniden hesapla
        sync_conn = db.get_conn()
        db.sync_rez_tahsilat(sync_conn, foy_no)
        sync_conn.commit(); sync_conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

@app.route('/api/tahsilat/adis', methods=['POST'])
def api_adis_tah():
    """Adisyon bazlı kısmi veya tam ödeme."""
    try:
        d = request.get_json()
        foy_no = int(d['foy_no'])
        odeme = d['odeme']
        adisyon_nolar = d.get('adisyon_nolar', [])  # seçili adisyon no listesi
        # Her adisyon için tutar ayrıca gönderilir: {adisyon_no: tutar}
        adis_tutarlar = d.get('adis_tutarlar', {})

        rez = db.get_rezervasyonlar()
        r = next((x for x in rez if x['foy_no'] == foy_no), None)
        otel = r.get('otel', cfg.otel_bilgi().get('kisa_ad','OTEL')) if r else 'LEO'
        musteri = r.get('musteri', '') if r else ''

        toplam_odeme = 0
        for adis_no in adisyon_nolar:
            tutar = float(adis_tutarlar.get(str(adis_no), 0))
            if tutar <= 0:
                continue
            toplam_odeme += tutar
            # Adisyon bazlı ödeme kaydet
            tarih = db.save_adisyon_odeme(adis_no, foy_no, tutar, odeme)

        if toplam_odeme > 0:
            # Rezervasyon tablosundaki adis_tahsilat güncelle
            db.save_adis_tahsilat(foy_no, toplam_odeme, odeme)

            # Yevmiye - her adisyon için 2 kayıt:
            # 1. hesap_kodu borç / 120 alacak (para geldi, müşteri borcu kapandı)
            # 2. 120 borç / 610 alacak (adisyon geliri gerçekleşti)
            hesap_kodu = ODEME_HESAP_KODU.get(odeme)
            if hesap_kodu:
                tarih = bugun().isoformat()
                conn = mdb.get_conn()
                for adis_no in adisyon_nolar:
                    t = float(adis_tutarlar.get(str(adis_no), 0))
                    if t > 0:
                        aciklama = f'Föy#{foy_no} Adis#{adis_no} {musteri} adisyon tahsilat'
                        # Para geldi: Kasa/Banka borç / Müşteri Cari alacak
                        mdb._yevmiye_ekle(conn, tarih, 'Adisyon Tahsilat - ' + odeme,
                                          hesap_kodu, '120', t, aciklama, otel)
                conn.commit(); conn.close()

        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400


@app.route('/api/tahsilat/adis-guncelle', methods=['POST'])
def api_adis_tah_guncelle():
    """Adisyon ödemesini direkt güncelle (sil+ekle değil)."""
    try:
        d = request.get_json()
        odeme_id = int(d['odeme_id'])   # adisyon_odemeler.id
        foy_no   = int(d['foy_no'])
        tutar    = float(d['tutar'])
        odeme    = d['odeme']
        tarih    = d.get('tarih') or bugun().isoformat()

        otel_conn = db.get_conn()

        # Eski tutarı al
        eski = otel_conn.execute(
            "SELECT tutar FROM adisyon_odemeler WHERE id=?", (odeme_id,)
        ).fetchone()
        eski_tutar = float(eski[0]) if eski else 0

        # adisyon_odemeler güncelle
        otel_conn.execute(
            "UPDATE adisyon_odemeler SET tutar=?, odeme_sekli=?, tarih=? WHERE id=?",
            (tutar, odeme, tarih, odeme_id)
        )

        # Rezervasyon adis_tahsilat farkını güncelle
        fark = tutar - eski_tutar
        otel_conn.execute("""
            UPDATE rezervasyonlar
            SET adis_tahsilat = adis_tahsilat + ?,
                adis_bakiye   = adis_bakiye   - ?
            WHERE foy_no=?
        """, (fark, fark, foy_no))

        db._sync_adisyon_totals(otel_conn, foy_no)
        otel_conn.commit(); otel_conn.close()

        # Yevmiyeyi de güncelle
        hesap_kodu = ODEME_HESAP_KODU.get(odeme, '102-1')
        conn = mdb.get_conn()
        # Eski yevmiye kaydını bul ve güncelle
        yev = conn.execute(
            "SELECT id FROM yevmiye WHERE islem_tipi LIKE 'Adisyon Tahsilat%' AND aciklama LIKE ? ORDER BY id DESC LIMIT 1",
            (f'%Adis%foy_no%',)
        ).fetchone()
        # Adisyon no üzerinden bul
        adis_row = db.get_conn().execute(
            "SELECT adisyon_no FROM adisyon_odemeler WHERE id=?", (odeme_id,)
        ).fetchone()
        if adis_row:
            adis_no = adis_row[0]
            conn.execute("""
                UPDATE yevmiye SET tutar=?, borc_hesap=?, tarih=?
                WHERE islem_tipi LIKE 'Adisyon Tahsilat%'
                AND aciklama LIKE ? ORDER BY id DESC LIMIT 1
            """, (tutar, hesap_kodu, tarih, f'%Adis#{adis_no}%'))
        conn.commit(); conn.close()

        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

@app.route('/api/adisyon/odemeler', methods=['GET'])
def api_adisyon_odemeler():
    """Bir adisyona ait ödeme geçmişi."""
    adisyon_no = request.args.get('adisyon_no', type=int)
    if not adisyon_no:
        return jsonify([])
    return jsonify(db.get_adisyon_odemeler(adisyon_no))

@app.route('/api/adisyon/ekle', methods=['POST'])
def api_adisyon_ekle():
    try:
        d = request.get_json()
        db.save_adisyon(d)
        # Yevmiye: Müşteri Cari borç / Adisyon Geliri alacak
        tutar = float(d.get('tutar') or 0)
        foy_no = int(d.get('foy_no') or 0)
        adisyon_no = int(d.get('adisyon_no') or 0)
        if tutar > 0 and foy_no:
            rez = db.get_rezervasyonlar()
            r = next((x for x in rez if x['foy_no'] == foy_no), None)
            otel = r.get('otel', 'GENEL') if r else 'GENEL'
            musteri = r.get('musteri', '') if r else ''
            tarih = d.get('tarih') or bugun().isoformat()
            conn = mdb.get_conn()
            mdb._yevmiye_ekle(conn, tarih, 'Adisyon Geliri', '120', '610',
                              tutar, f'Föy#{foy_no} Adis#{adisyon_no} {musteri} adisyon', otel)
            conn.commit(); conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

@app.route('/api/adisyon/guncelle', methods=['POST'])
def api_adisyon_guncelle():
    try:
        d = request.get_json()
        adisyon_no = int(d['adisyon_no'])
        yeni_tutar = float(d['tutar'])
        odeme = d.get('odeme', 'Oda Hesabına')

        # Eski adisyon bilgisini al
        conn_db = db.get_conn()
        a = conn_db.execute("SELECT * FROM adisyonlar WHERE adisyon_no=?", (adisyon_no,)).fetchone()
        conn_db.close()

        if a:
            eski_tutar = float(a['tutar'] or 0)
            foy_no = a['foy_no']
            musteri_map = {r['foy_no']: r['musteri'] for r in db.get_rezervasyonlar()}
            musteri = musteri_map.get(foy_no, '')

            # Yevmiyedeki Adisyon Geliri kaydını güncelle
            conn_muh = mdb.get_conn()
            conn_muh.execute(
                "UPDATE yevmiye SET tutar=? WHERE islem_tipi='Adisyon Geliri' AND aciklama LIKE ?",
                (yeni_tutar, f'Föy#{foy_no} Adis#{adisyon_no}%')
            )
            conn_muh.execute(
                "UPDATE yevmiye SET tutar=? WHERE islem_tipi='Adisyon Geliri' AND aciklama=?",
                (yeni_tutar, f'Föy#{foy_no} {musteri} adisyon')
            )
            conn_muh.commit(); conn_muh.close()

        # Adisyonu güncelle (tarih, oda ve açıklama da değişebilir)
        yeni_tarih = d.get('tarih')
        yeni_foy   = int(d['foy_no']) if d.get('foy_no') else None
        yeni_aciklama = d.get('aciklama')
        db.update_adisyon(adisyon_no, yeni_tutar, odeme, tarih=yeni_tarih, foy_no=yeni_foy, aciklama=yeni_aciklama)

        # Rezervasyon adisyon toplamını yeniden hesapla (fark yerine toplam)
        if a:
            foy_no = a['foy_no']
            conn_db2 = db.get_conn()
            toplam = conn_db2.execute(
                "SELECT COALESCE(SUM(tutar),0) FROM adisyonlar WHERE foy_no=?", (foy_no,)
            ).fetchone()[0]
            tah = conn_db2.execute(
                "SELECT COALESCE(SUM(tutar),0) FROM adisyon_odemeler WHERE foy_no=?", (foy_no,)
            ).fetchone()[0]
            conn_db2.execute(
                "UPDATE rezervasyonlar SET adisyon=?, adis_bakiye=? WHERE foy_no=?",
                (toplam, toplam - tah, foy_no)
            )
            conn_db2.commit(); conn_db2.close()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

@app.route('/api/adisyon/sil', methods=['POST'])
def api_adisyon_sil():
    try:
        adisyon_no = int(request.get_json()['adisyon_no'])
        # Adisyon bilgisini al
        conn_db = db.get_conn()
        a = conn_db.execute("SELECT * FROM adisyonlar WHERE adisyon_no=?", (adisyon_no,)).fetchone()
        conn_db.close()
        if a:
            a = dict(a)
            foy_no = a['foy_no']
            tutar = float(a['tutar'] or 0)
            otel = a.get('otel', 'GENEL')
            musteri_map = {r['foy_no']: r['musteri'] for r in db.get_rezervasyonlar()}
            musteri = musteri_map.get(foy_no, '')
            # Yevmiyeden Adisyon Geliri ve Tahsilat kayıtlarını sil
            conn_muh = mdb.get_conn()
            conn_muh.execute(
                "DELETE FROM yevmiye WHERE aciklama LIKE ? AND islem_tipi='Adisyon Geliri'",
                (f'Föy#{foy_no}%Adis#{adisyon_no}%',)
            )
            # Adis no olmayan eski kayıtlar için de dene
            conn_muh.execute(
                "DELETE FROM yevmiye WHERE aciklama=? AND islem_tipi='Adisyon Geliri'",
                (f'Föy#{foy_no} {musteri} adisyon',)
            )
            # Adisyon tahsilat kayıtlarını da sil
            conn_muh.execute(
                "DELETE FROM yevmiye WHERE aciklama LIKE ? AND islem_tipi LIKE 'Adisyon Tahsilat%'",
                (f'Föy#{foy_no} Adis#{adisyon_no}%',)
            )
            conn_muh.commit(); conn_muh.close()
            # adisyon_odemeler tablosundan sil
            conn_db2 = db.get_conn()
            conn_db2.execute("DELETE FROM adisyon_odemeler WHERE adisyon_no=?", (adisyon_no,))
            conn_db2.commit(); conn_db2.close()
            # Rezervasyon adisyon toplamını güncelle
            rez_list = db.get_rezervasyonlar()
            r = next((x for x in rez_list if x['foy_no'] == foy_no), None)
            if r:
                odenen = float(a['odenen_tutar'] or 0)
                conn_db3 = db.get_conn()
                conn_db3.execute(
                    "UPDATE rezervasyonlar SET adisyon=adisyon-?, adis_tahsilat=adis_tahsilat-?, adis_bakiye=adis_bakiye-? WHERE foy_no=?",
                    (tutar, odenen, tutar - odenen, foy_no)
                )
                conn_db3.commit(); conn_db3.close()
        # Adisyonu sil
        db.delete_adisyon(adisyon_no)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

@app.route('/yedek/db')
@login_required
def yedek_db():
    """Hem otel hem muhasebe veritabanlarını ZIP olarak indir."""
    import shutil, tempfile, zipfile, os
    from datetime import date
    dosya_adi = f"yedek_{bugun().isoformat()}.zip"
    tmp_zip = tempfile.NamedTemporaryFile(suffix='.zip', delete=False)
    tmp_zip.close()
    with zipfile.ZipFile(tmp_zip.name, 'w', zipfile.ZIP_DEFLATED) as zf:
        otel_db = str(db.DB_PATH)
        muh_db  = str(mdb.DB_PATH)
        if os.path.exists(otel_db):
            zf.write(otel_db, 'otel.db')
        if os.path.exists(muh_db):
            zf.write(muh_db, 'muhasebe.db')
    return send_file(tmp_zip.name, as_attachment=True, download_name=dosya_adi,
                     mimetype='application/zip')

def excel_yedek_olustur():
    """Tüm DB'yi Excel olarak üretir, geçici dosya yolunu döner."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    HEADER_FONT  = Font(bold=True, color='FFFFFF')
    HEADER_FILL  = PatternFill('solid', fgColor='0d1b2a')
    HEADER2_FILL = PatternFill('solid', fgColor='1a2f4a')

    def make_header(ws, headers, fill=None):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = HEADER_FONT
            c.fill = fill or HEADER_FILL
        for col in ws.columns:
            ml = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(ml+4, 35)

    # 1. Rezervasyon Girişleri
    ws1 = wb.active
    ws1.title = 'Rezervasyon Girisleri'
    h1 = ['Oda No','Otel','Foy No','Acente','Musteri Adi','Yetiskin','Cocuk',
          'Ek Yatak','Gun Fiyat','Giris Tarihi','Cikis Tarihi','Toplam Gun',
          'Toplam Fiyat','Kapora','Kapora Tarihi','Tahsilat','Odeme Turu',
          'REZ Bakiye','Adisyon','Adisyon Tahsilat','Odeme Turu2','ADS Bakiye','Aciklama']
    make_header(ws1, h1)
    for i, r in enumerate(db.get_rezervasyonlar(), 2):
        vals = [r.get('oda_no'), r.get('otel'), r.get('foy_no'), r.get('kanal'),
                r.get('musteri'), r.get('yetiskin'), r.get('cocuk'), r.get('ek_yatak'),
                r.get('gun_fiyat'), r.get('giris'), r.get('cikis'), r.get('toplam_gun'),
                r.get('toplam_fiyat'), r.get('kapora'), r.get('kapora_tarihi'),
                r.get('rez_tahsilat'), r.get('rez_odeme_sekli'), r.get('rez_bakiye'),
                r.get('adisyon'), r.get('adis_tahsilat'), r.get('adis_odeme_sekli'),
                r.get('adis_bakiye'), r.get('aciklama')]
        for col, v in enumerate(vals, 1):
            ws1.cell(row=i, column=col, value=v)

    # 2. Adisyonlar
    ws2 = wb.create_sheet('Adisyonlar')
    h2 = ['Adisyon No','Foy No','Oda No','Tarih','Tutar','Odeme','Otel','Aciklama']
    make_header(ws2, h2)
    for i, a in enumerate(db.get_adisyonlar(), 2):
        vals = [a.get('adisyon_no'), a.get('foy_no'), a.get('oda_no'),
                a.get('tarih'), a.get('tutar'), a.get('odeme'),
                a.get('otel'), a.get('aciklama')]
        for col, v in enumerate(vals, 1):
            ws2.cell(row=i, column=col, value=v)

    conn = mdb.get_conn()

    # 3. Yevmiye
    ws3 = wb.create_sheet('YEVMIYE')
    h3 = ['BELGE NO','TARIH','ISLEM TIPI','BORC HESABI','ALACAK HESABI','TUTAR','ACIKLAMA','FATURA NO']
    make_header(ws3, h3, HEADER2_FILL)
    yevmiye = [dict(r) for r in conn.execute("SELECT * FROM yevmiye ORDER BY tarih, id").fetchall()]
    for i, r in enumerate(yevmiye, 2):
        vals = [r.get('belge_no'), r.get('tarih'), r.get('islem_tipi'),
                r.get('borc_hesap'), r.get('alacak_hesap'),
                r.get('tutar'), r.get('aciklama'), r.get('fatura_no')]
        for col, v in enumerate(vals, 1):
            ws3.cell(row=i, column=col, value=v)

    # 4. Personel
    ws4 = wb.create_sheet('PERSONEL')
    h4 = ['Ad Soyad','Ise Giris','Gorev','Net Maas','Banka IBAN','TC Kimlik','Telefon','Aktif']
    make_header(ws4, h4, HEADER2_FILL)
    personel = mdb.get_personel()
    for i, p in enumerate(personel, 2):
        vals = [p.get('ad_soyad'), p.get('ise_giris'), p.get('gorev'),
                p.get('net_maas'), p.get('banka_iban'), p.get('tc_kimlik'),
                p.get('telefon'), p.get('aktif')]
        for col, v in enumerate(vals, 1):
            ws4.cell(row=i, column=col, value=v)

    # 4b. Maaş ödemeleri
    ws4b = wb.create_sheet('PERSONEL MAAS')
    h4b = ['Personel','Tarih','Donem Yil','Donem Ay','Net Odeme','Yol','Mesai','Izin','Avans Dusum','Odeme Banka','Aciklama']
    make_header(ws4b, h4b, HEADER2_FILL)
    maaslar = [dict(r) for r in conn.execute(
        "SELECT pm.*, p.ad_soyad FROM personel_maas pm JOIN personel p ON pm.personel_id=p.id ORDER BY pm.tarih"
    ).fetchall()]
    for i, m in enumerate(maaslar, 2):
        vals = [m.get('ad_soyad'), m.get('tarih'), m.get('donem_yil'), m.get('donem_ay'),
                m.get('net_odeme'), m.get('yol_parasi'), m.get('fazla_mesai'),
                m.get('izin_parasi'), m.get('avans_dusum'), m.get('odeme_banka'), m.get('aciklama')]
        for col, v in enumerate(vals, 1):
            ws4b.cell(row=i, column=col, value=v)

    # 4c. Avans ödemeleri
    ws4c = wb.create_sheet('PERSONEL AVANS')
    h4c = ['Personel','Tarih','Tutar','Odeme Sekli','Aciklama']
    make_header(ws4c, h4c, HEADER2_FILL)
    avanslar = [dict(r) for r in conn.execute(
        "SELECT a.*, p.ad_soyad FROM personel_avans a JOIN personel p ON p.id=a.personel_id ORDER BY a.tarih"
    ).fetchall()]
    for i, a in enumerate(avanslar, 2):
        vals = [a.get('ad_soyad'), a.get('tarih'), a.get('tutar'),
                a.get('odeme_sekli'), a.get('aciklama')]
        for col, v in enumerate(vals, 1):
            ws4c.cell(row=i, column=col, value=v)

    # 5. Stok
    ws5 = wb.create_sheet('STOK')
    h5 = ['TARIH','BELGE NO','ACIKLAMA','KATEGORI','TUTAR','ODEME HESABI','FATURA','NOT']
    make_header(ws5, h5, HEADER2_FILL)
    stok = [dict(r) for r in conn.execute("SELECT * FROM stok ORDER BY tarih").fetchall()]
    for i, r in enumerate(stok, 2):
        vals = [r.get('tarih'), r.get('belge_no'), r.get('aciklama'), r.get('kategori'),
                r.get('tutar'), r.get('odeme_hesap'), r.get('fatura_var'), r.get('not_')]
        for col, v in enumerate(vals, 1):
            ws5.cell(row=i, column=col, value=v)

    # 5b. Demirbaş
    ws5b = wb.create_sheet('DEMIRBAS')
    h5b = ['TARIH','ACIKLAMA','MIKTAR','BIRIM FIYAT','TOPLAM','ODEME HESABI','FATURA NO','NOT']
    make_header(ws5b, h5b, HEADER2_FILL)
    demirbas = [dict(r) for r in conn.execute("SELECT * FROM \"demirbaş\" ORDER BY tarih").fetchall()]
    for i, r in enumerate(demirbas, 2):
        vals = [r.get('tarih'), r.get('aciklama'), r.get('miktar'), r.get('birim_fiyat'),
                r.get('toplam'), r.get('odeme_hesap'), r.get('fatura_no'), r.get('not_')]
        for col, v in enumerate(vals, 1):
            ws5b.cell(row=i, column=col, value=v)

    # 6. Vergi
    ws6 = wb.create_sheet('VERGI')
    h6 = ['TARIH','DONEM YIL','DONEM AY','VERGI TURU','MATRAH','TUTAR','ODEME BANKASI','DURUM','ACIKLAMA']
    make_header(ws6, h6, HEADER2_FILL)
    vergi = [dict(r) for r in conn.execute("SELECT * FROM vergi ORDER BY donem_yil, donem_ay").fetchall()]
    for i, r in enumerate(vergi, 2):
        vals = [r.get('tarih'), r.get('donem_yil'), r.get('donem_ay'), r.get('vergi_turu'),
                r.get('matrah'), r.get('tutar'), r.get('odeme_banka'), r.get('durum'), r.get('aciklama')]
        for col, v in enumerate(vals, 1):
            ws6.cell(row=i, column=col, value=v)

    # 7. Acente Cari
    ws7 = wb.create_sheet('ACENTE CARI')
    h7 = ['TARIH','ACENTE','FOY NO','REZ NO','MISAFIR','REZ TUTARI','KOM ORAN','KOMISYON TL','GELEN ODEME','OTEL']
    make_header(ws7, h7, HEADER2_FILL)
    acente = [dict(r) for r in conn.execute("SELECT * FROM acente_cari ORDER BY tarih").fetchall()]
    for i, r in enumerate(acente, 2):
        vals = [r.get('tarih'), r.get('acente_kod'), r.get('foy_no'), r.get('rez_no'),
                r.get('misafir'), r.get('rez_tutari'), r.get('komisyon_oran'),
                r.get('komisyon_tl'), r.get('gelen_odeme'), r.get('otel')]
        for col, v in enumerate(vals, 1):
            ws7.cell(row=i, column=col, value=v)

    # 8. Ortak Cari
    ws8 = wb.create_sheet('ORTAK CARI')
    h8 = ['TARIH','ORTAK','BELGE NO','ACIKLAMA','KATEGORI','TUTAR','ODEME','IADE','NET','OTEL']
    make_header(ws8, h8, HEADER2_FILL)
    ortak = [dict(r) for r in conn.execute("SELECT * FROM ortak_cari ORDER BY tarih").fetchall()]
    for i, r in enumerate(ortak, 2):
        net = (r.get('tutar') or 0) - (r.get('iade') or 0)
        vals = [r.get('tarih'), r.get('ortak'), r.get('belge_no'), r.get('aciklama'),
                r.get('gider_kategori'), r.get('tutar'), r.get('odeme_sekli'),
                r.get('iade'), net, r.get('otel')]
        for col, v in enumerate(vals, 1):
            ws8.cell(row=i, column=col, value=v)

    conn.close()

    # Sütun genişlikleri
    for ws in wb.worksheets:
        for col in ws.columns:
            ml = max((len(str(c.value or '')) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(ml+4, 40)

    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp.close()
    wb.save(tmp.name)
    return tmp.name


def yedek_mail_gonder():
    """Excel yedeğini oluşturup yedek mail adresine gönderir."""
    gmail_user = os.environ.get('GMAIL_USER', 'bmkucuk@gmail.com')
    gmail_pass = os.environ.get('GMAIL_APP_PASSWORD', '')
    alici      = cfg.get('sistem.yedek_mail_alici', gmail_user)
    if not gmail_pass:
        print('[YEDEK] GMAIL_APP_PASSWORD tanımlı değil, mail atlanıyor.')
        return
    try:
        dosya_yolu = excel_yedek_olustur()
        dosya_adi  = f"otel_yedek_{bugun().isoformat()}.xlsx"

        msg = MIMEMultipart()
        msg['From']    = gmail_user
        msg['To']      = alici
        msg['Subject'] = f"🏨 Otel Yönetim Günlük Yedek — {bugun().strftime('%d.%m.%Y')}"
        msg.attach(MIMEText(
            f"Merhaba,\n\n{cfg.otel_bilgi().get('ad', 'Otel')} yönetim sistemi günlük yedeği ektedir.\n"
            f"Tarih: {bugun().strftime('%d.%m.%Y')}\n\n"
            f"Bu mail otomatik olarak gönderilmiştir.", 'plain', 'utf-8'))

        with open(dosya_yolu, 'rb') as f:
            ek = MIMEApplication(f.read(), Name=dosya_adi)
            ek['Content-Disposition'] = f'attachment; filename="{dosya_adi}"'
            msg.attach(ek)

        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
            smtp.login(gmail_user, gmail_pass)
            smtp.send_message(msg)

        os.unlink(dosya_yolu)
        print(f'[YEDEK] Mail başarıyla gönderildi: {dosya_adi}')
    except Exception as e:
        print(f'[YEDEK] Mail gönderilemedi: {e}')


@app.route('/yedek/excel')
@login_required
def yedek_excel():
    dosya_yolu = excel_yedek_olustur()
    dosya_adi  = f"otel_yedek_{bugun().isoformat()}.xlsx"
    return send_file(dosya_yolu, as_attachment=True, download_name=dosya_adi,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/geri-yukle', methods=['GET'])
@admin_required
def geri_yukle_sayfa():
    return render_template('geri_yukle.html')


@app.route('/api/geri-yukle', methods=['POST'])
@admin_required
def api_geri_yukle():
    """Yedek Excel'den tüm tabloları geri yükler."""
    try:
        from openpyxl import load_workbook
        f = request.files.get('excel')
        if not f:
            return jsonify({'ok': False, 'error': 'Dosya seçilmedi'}), 400

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            f.save(tmp.name)
            tmp_path = tmp.name

        wb = load_workbook(tmp_path, data_only=True)
        os.unlink(tmp_path)
        ozet = {}

        conn_m = mdb.get_conn()
        conn_o = db.get_conn()

        def col(row, idx):
            v = row[idx]
            return v if v is not None else None

        # STOK
        if 'STOK' in wb.sheetnames:
            ws = wb['STOK']
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            rows = [r for r in rows if any(v for v in r)]
            conn_m.execute("DELETE FROM stok")
            for r in rows:
                conn_m.execute(
                    "INSERT INTO stok(tarih,belge_no,aciklama,kategori,tutar,odeme_hesap,fatura_var,not_) VALUES(?,?,?,?,?,?,?,?)",
                    (col(r,0), col(r,1), col(r,2) or '', col(r,3), col(r,4) or 0, col(r,5), col(r,6), col(r,7)))
            ozet['stok'] = len(rows)

        # DEMİRBAŞ
        if 'DEMIRBAS' in wb.sheetnames:
            ws = wb['DEMIRBAS']
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            rows = [r for r in rows if any(v for v in r)]
            conn_m.execute('DELETE FROM "demirbaş"')
            for r in rows:
                conn_m.execute(
                    'INSERT INTO "demirbaş"(tarih,aciklama,miktar,birim_fiyat,toplam,odeme_hesap,fatura_no,not_) VALUES(?,?,?,?,?,?,?,?)',
                    (col(r,0), col(r,1) or '', col(r,2) or 1, col(r,3) or 0, col(r,4), col(r,5), col(r,6), col(r,7)))
            ozet['demirbas'] = len(rows)

        # YEVMİYE
        if 'YEVMIYE' in wb.sheetnames:
            ws = wb['YEVMIYE']
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            rows = [r for r in rows if any(v for v in r)]
            conn_m.execute("DELETE FROM yevmiye")
            for r in rows:
                tarih = str(col(r,1) or '')
                yil = int(tarih[:4]) if len(tarih) >= 4 else None
                ay  = int(tarih[5:7]) if len(tarih) >= 7 else None
                conn_m.execute(
                    "INSERT INTO yevmiye(belge_no,tarih,islem_tipi,borc_hesap,alacak_hesap,tutar,aciklama,fatura_no,yil,ay,otel) VALUES(?,?,?,?,?,?,?,?,?,?,'GENEL')",
                    (col(r,0), tarih, col(r,2) or '', col(r,3) or '', col(r,4) or '', col(r,5) or 0, col(r,6), col(r,7), yil, ay))
            ozet['yevmiye'] = len(rows)

        # PERSONEL (önce personel yüklenmeli ki maaş/avans ID'leri eşleşsin)
        if 'PERSONEL' in wb.sheetnames:
            ws = wb['PERSONEL']
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            rows = [r for r in rows if any(v for v in r)]
            conn_m.execute("DELETE FROM personel")
            for r in rows:
                conn_m.execute(
                    "INSERT INTO personel(ad_soyad,ise_giris,gorev,net_maas,banka_iban,tc_kimlik,telefon,aktif) VALUES(?,?,?,?,?,?,?,?)",
                    (col(r,0) or '', col(r,1), col(r,2), col(r,3) or 0,
                     col(r,4), col(r,5), col(r,6), col(r,7) if col(r,7) is not None else 1))
            ozet['personel'] = len(rows)

        # PERSONEL MAAS
        if 'PERSONEL MAAS' in wb.sheetnames:
            ws = wb['PERSONEL MAAS']
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            rows = [r for r in rows if any(v for v in r)]
            conn_m.execute("DELETE FROM personel_maas")
            personel_liste = {p['ad_soyad']: p['id'] for p in mdb.get_personel()}
            for r in rows:
                pid = personel_liste.get(col(r,0))
                if not pid:
                    continue
                conn_m.execute(
                    "INSERT INTO personel_maas(personel_id,tarih,donem_yil,donem_ay,net_odeme,yol_parasi,fazla_mesai,izin_parasi,avans_dusum,odeme_banka,aciklama) VALUES(?,?,?,?,?,?,?,?,?,?,?)",
                    (pid, col(r,1), col(r,2), col(r,3), col(r,4) or 0, col(r,5) or 0, col(r,6) or 0, col(r,7) or 0, col(r,8) or 0, col(r,9), col(r,10)))
            ozet['personel_maas'] = len(rows)

        # PERSONEL AVANS
        if 'PERSONEL AVANS' in wb.sheetnames:
            ws = wb['PERSONEL AVANS']
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            rows = [r for r in rows if any(v for v in r)]
            conn_m.execute("DELETE FROM personel_avans")
            personel_liste = {p['ad_soyad']: p['id'] for p in mdb.get_personel()}
            for r in rows:
                pid = personel_liste.get(col(r,0))
                if not pid:
                    continue
                conn_m.execute(
                    "INSERT INTO personel_avans(personel_id,tarih,tutar,odeme_sekli,aciklama) VALUES(?,?,?,?,?)",
                    (pid, col(r,1), col(r,2) or 0, col(r,3), col(r,4)))
            ozet['personel_avans'] = len(rows)

        # REZERVASYONlar
        if 'Rezervasyon Girisleri' in wb.sheetnames:
            rez_count, adis_count = 0, 0
            # Mevcut import fonksiyonunu kullan
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp2:
                wb.save(tmp2.name)
                tmp2_path = tmp2.name
            try:
                rez_count, adis_count = db.import_from_excel(tmp2_path)
            finally:
                os.unlink(tmp2_path)
            ozet['rezervasyon'] = rez_count
            ozet['adisyon'] = adis_count

        conn_m.commit()
        conn_m.close()
        conn_o.close()

        return jsonify({'ok': True, 'ozet': ozet})
    except Exception as e:
        import traceback
        return jsonify({'ok': False, 'error': str(e), 'detail': traceback.format_exc()}), 400




# ── Scheduler: Her gece 02:00 otomatik yedek maili ───────────────────────────
_scheduler = BackgroundScheduler(timezone=TR_TZ)
_scheduler.add_job(yedek_mail_gonder, 'cron', hour=2, minute=0)
_scheduler.start()



# Tek seferlik migration — fonksiyon tanımlarının hepsi yüklendikten sonra çalışır
migrate_acente_otomatik()

# Eski tasarımdan kalan, artık hiçbir sayfada kullanılmayan föy bazlı acente_cari
# kayıtlarını (Telefon/Kapıdan/EXP gibi kanallar için otomatik oluşturulmuş, hep
# 0 ₺ görünen satırlar) temizle. Yeni kod hiçbir zaman böyle satır oluşturmuyor,
# bu yüzden her başlangıçta çalıştırmak güvenli ve ucuz.
try:
    _c = mdb.get_conn()
    _c.execute("DELETE FROM acente_cari WHERE foy_no IS NOT NULL")
    _c.commit(); _c.close()
except Exception as _e:
    print(f'Eski acente_cari temizliği hatası: {_e}')

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    print(f"🌐 Sunucu → http://localhost:{port}")
    app.run(debug=False, host='0.0.0.0', port=port)


# ── Tema ────────────────────────────────────────────────────────────────────

@app.route('/api/tema-kaydet', methods=['POST'])
@login_required
def api_tema_kaydet():
    data = request.get_json()
    mod  = data.get('mod', 'dark')
    if mod not in ('dark', 'light'):
        return jsonify({'ok': False})
    c = cfg.load_config()
    c['tema']['mod'] = mod
    cfg.save_config(c)
    return jsonify({'ok': True, 'mod': mod})




# ── Kurulum Sihirbazı ─────────────────────────────────────────────────────────

def kurulum_tamamlandi_mi():
    """config.json'da otel adı girilmişse kurulum tamamlanmış demektir."""
    return cfg.get('otel.ad', 'Otel Adı') != 'Otel Adı'

@app.route('/kurulum')
def kurulum_sayfasi():
    if kurulum_tamamlandi_mi():
        return redirect('/login')
    return render_template('kurulum.html')

@app.route('/kurulum/telegram-test', methods=['POST'])
def kurulum_telegram_test():
    """Token ile getUpdates yapıp chat_id bulur."""
    import urllib.request, json as _json
    data  = request.get_json()
    token = data.get('token', '').strip()
    if not token:
        return jsonify({'ok': False, 'hata': 'Token boş'})
    try:
        url = f'https://api.telegram.org/bot{token}/getUpdates'
        req = urllib.request.Request(url)
        with urllib.request.urlopen(req, timeout=8) as r:
            resp = _json.loads(r.read())
        if not resp.get('ok'):
            return jsonify({'ok': False, 'hata': 'Geçersiz token'})
        # Chat ID bul
        for update in reversed(resp.get('result', [])):
            chat = update.get('message', {}).get('chat', {})
            if chat.get('type') in ('group', 'supergroup'):
                return jsonify({'ok': True, 'chat_id': str(chat['id']),
                                'chat_adi': chat.get('title', '')})
        return jsonify({'ok': False,
                        'hata': 'Grup mesajı bulunamadı. Botu gruba ekleyip bir mesaj yazın.'})
    except Exception as e:
        return jsonify({'ok': False, 'hata': str(e)})

@app.route('/kurulum/kaydet', methods=['POST'])
def kurulum_kaydet():
    """Sihirbazdan gelen veriyi config.json ve DB'ye yazar."""
    import hashlib
    from datetime import date as _date

    data  = request.get_json()
    otel  = data.get('otel', {})
    odalar = data.get('odalar', {})
    kullanicilar = data.get('kullanicilar', [])
    enteg = data.get('entegrasyon', {})

    if not otel.get('ad') or not otel.get('kisa_ad'):
        return jsonify({'ok': False, 'hata': 'Otel adı ve kısa kod zorunludur'})

    # 1. config.json güncelle
    c = cfg.load_config()
    c['otel']['ad']            = otel.get('ad', '')
    c['otel']['kisa_ad']       = otel.get('kisa_ad', '').upper()
    c['otel']['sehir']         = otel.get('sehir', '')
    c['otel']['telefon']       = otel.get('telefon', '')
    c['otel']['toplam_oda']    = int(odalar.get('sayi', 20))
    c['otel']['oda_baslangic'] = int(odalar.get('bas', 1))
    c['otel']['oda_bitis']     = int(odalar.get('bitis', 20))
    c['sistem']['yedek_mail_alici']  = enteg.get('yedek_mail', '')
    c['sistem']['telegram_token']    = enteg.get('tg_token', '')
    c['sistem']['telegram_chat_id']  = enteg.get('tg_chat', '')
    c['sistem']['demo_baslangic']    = _date.today().isoformat()
    c['sistem']['demo_sure_gun']     = 3
    c['sistem']['lisans_aktif']      = False
    c['sistem']['askiya_alindi']     = False

    # Partner bilgilerini config'e yaz
    ortaklar = []
    for u in kullanicilar:
        if u.get('rol') == 'partner':
            ortaklar.append({
                'kod':    u.get('kisalt', 'P' + str(u.get('idx', 1))),
                'kisalt': u.get('kisalt', 'P' + str(u.get('idx', 1))),
                'ad':     u.get('adsoyad', u.get('kullanici', ''))
            })
    c['ortaklar'] = ortaklar
    cfg.save_config(c)

    # 2. Kullanıcıları DB'ye yaz
    conn = db.get_conn()
    conn.execute("DELETE FROM kullanicilar WHERE username != 'superadmin'")
    for u in kullanicilar:
        h = hashlib.sha256(u['sifre'].encode()).hexdigest()
        rol = u.get('rol', 'resepsiyon')
        conn.execute(
            "INSERT OR REPLACE INTO kullanicilar(username, ad, hash, role) VALUES(?,?,?,?)",
            (u['kullanici'], u['kullanici'], h, rol)
        )
    conn.commit()
    conn.close()

    # 3. TELEGRAM_TOKEN env'e yaz (runtime için)
    if enteg.get('tg_token'):
        os.environ['TELEGRAM_TOKEN']   = enteg['tg_token']
        os.environ['TELEGRAM_CHAT_ID'] = enteg.get('tg_chat', '')

    return jsonify({'ok': True})

# ── Lisans Sayfaları ──────────────────────────────────────────────────────────

@app.route('/askida')
def sayfa_askida():
    return render_template('lisans/askida.html')

@app.route('/demo-bitti')
def sayfa_demo_bitti():
    return render_template('lisans/demo_bitti.html')

# ── Süper Admin API (sadece SUPERADMIN_KEY ile erişilebilir) ─────────────────

SUPERADMIN_KEY = os.environ.get('SUPERADMIN_KEY', '')

def superadmin_kontrol():
    key = request.headers.get('X-Admin-Key', '') or request.args.get('key', '')
    return key and key == SUPERADMIN_KEY

@app.route('/sadmin/aktif', methods=['POST'])
def sadmin_aktif():
    if not superadmin_kontrol():
        return jsonify({'ok': False, 'error': 'Yetkisiz'}), 403
    c = cfg.load_config()
    c['sistem']['askiya_alindi'] = False
    c['sistem']['lisans_aktif']  = True
    cfg.save_config(c)
    return jsonify({'ok': True, 'durum': 'aktif'})

@app.route('/sadmin/askiya-al', methods=['POST'])
def sadmin_askiya_al():
    if not superadmin_kontrol():
        return jsonify({'ok': False, 'error': 'Yetkisiz'}), 403
    c = cfg.load_config()
    c['sistem']['askiya_alindi'] = True
    cfg.save_config(c)
    return jsonify({'ok': True, 'durum': 'askida'})

@app.route('/sadmin/demo-uzat', methods=['POST'])
def sadmin_demo_uzat():
    if not superadmin_kontrol():
        return jsonify({'ok': False, 'error': 'Yetkisiz'}), 403
    data = request.get_json() or {}
    gun  = int(data.get('gun', 3))
    c    = cfg.load_config()
    from datetime import date
    c['sistem']['demo_baslangic'] = date.today().isoformat()
    c['sistem']['demo_sure_gun']  = gun
    c['sistem']['lisans_aktif']   = False
    c['sistem']['askiya_alindi']  = False
    cfg.save_config(c)
    return jsonify({'ok': True, 'gun': gun})

@app.route('/sadmin')
def sadmin_sayfa():
    return render_template('sadmin.html')

@app.route('/sadmin/durum')
def sadmin_durum():
    if not superadmin_kontrol():
        return jsonify({'ok': False, 'error': 'Yetkisiz'}), 403
    c = cfg.load_config()
    return jsonify({
        'ok':     True,
        'durum':  cfg.lisans_durumu(),
        'kalan':  cfg.demo_kalan_gun(),
        'config': c
    })

@app.route('/sadmin/not-kaydet', methods=['POST'])
def sadmin_not_kaydet():
    if not superadmin_kontrol():
        return jsonify({'ok': False, 'error': 'Yetkisiz'}), 403
    data = request.get_json()
    c = cfg.load_config()
    c['sistem']['not'] = data.get('not', '')
    cfg.save_config(c)
    return jsonify({'ok': True})
